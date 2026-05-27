from __future__ import annotations

import csv
import json
import re
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from backend.config import Config
from backend.services.llm_client import get_llm_client, _extract_json
from backend.services.validators import normalize_currency, normalize_date


_RESERVED_FIELD_KEYS = {
    "classification",
    "document_id",
    "document_type",
    "entities",
    "field_annotations",
    "fields",
    "layout_context",
    "metadata",
    "ocr",
    "pages",
    "requires_review",
    "status",
    "summary",
    "validation",
}

_TEXT_ONLY_FIELD_KEYS = {"text", "raw_text", "raw_text_source"}


def build_document_payload(document: Any, extraction_results: list[dict] | None = None, tables: list[dict] | None = None) -> dict:
    # Produce a clean frontend-ready payload that hides OCR internals.
    structured = getattr(document, "extraction_result", None) or {}

    field_map, text_lines = _build_field_payload(structured, extraction_results)
    field_rows = _field_rows(field_map, extraction_results)

    # summary / entities / tables live in top-level structured keys or as underscored keys
    summary = structured.get("summary") or structured.get("_summary") or ""
    entities = structured.get("entities") or structured.get("_entities") or {}
    tables = _normalize_tables(structured, tables)
    text_block = "\n".join(text_lines).strip()

    metadata = structured.get("metadata") or {"detected_language": "", "page_count": getattr(document, "page_count", 1)}

    return {
        "document_id": getattr(document, "id", None),
        "document_type": structured.get("classification", {}).get("type") or getattr(document, "doc_type", None) or structured.get("document_type"),
        "summary": summary,
        "entities": entities,
        "fields": field_map,
        "field_rows": field_rows,
        "text": text_block,
        "tables": tables,
        "metadata": metadata,
        "status": getattr(document, "status", None),
    }


def build_export_payload(document: Any, extraction_results: list[dict] | None = None, tables: list[dict] | None = None) -> dict:
    structured = getattr(document, "extraction_result", None) or {}
    extraction_results = extraction_results or []
    tables = tables or []

    inferred_type = _infer_export_document_type(document, structured, extraction_results)
    if inferred_type == "invoice":
        return _build_invoice_export_payload(document, structured, extraction_results, tables)
    if inferred_type == "receipt":
        return _build_receipt_export_payload(document, structured, extraction_results, tables)
    if inferred_type == "business_card":
        return _build_business_card_export_payload(document, structured, extraction_results, tables)
    if inferred_type == "bank_statement":
        return _build_bank_statement_export_payload(document, structured, extraction_results, tables)

    return build_document_payload(document, extraction_results, tables)


def export_json_file(payload: dict, output_path: str) -> str:
    path = Path(output_path)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return str(path)


def export_csv_file(fields: list[dict], output_path: str) -> str:
    path = Path(output_path)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["field_name", "field_value", "confidence"])
        writer.writeheader()
        for item in fields:
            writer.writerow({"field_name": item.get("field_name"), "field_value": item.get("field_value"), "confidence": item.get("confidence")})
    return str(path)


def export_excel_file(payload: dict, output_path: str) -> str:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Field", "Value"])
    for key, value in payload.get("document", {}).items():
        summary_sheet.append([key, value])

    fields_sheet = workbook.create_sheet("Fields")
    fields_sheet.append(["Field", "Value", "Confidence"])
    for item in payload.get("fields", []):
        fields_sheet.append([item.get("field_name"), item.get("field_value"), item.get("confidence")])

    tables_sheet = workbook.create_sheet("Tables")
    tables_sheet.append(["Table Index", "Headers", "Rows"])
    for index, table in enumerate(payload.get("tables", []), start=1):
        tables_sheet.append([index, json.dumps(table.get("headers", [])), json.dumps(table.get("rows", []))])

    workbook.save(output_path)
    return output_path


def _field_lookup(structured: dict[str, Any], extraction_results: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    field_map, _ = _build_field_payload(structured, extraction_results)
    return field_map


def _build_field_payload(structured: dict[str, Any], extraction_results: list[dict[str, Any]] | None = None) -> tuple[dict[str, Any], list[str]]:
    field_map: dict[str, Any] = {}
    text_lines: list[str] = []

    def add_field(name: Any, value: Any) -> None:
        if not name:
            return
        normalized_name = str(name).strip()
        if not normalized_name or normalized_name.startswith("_") or normalized_name in _RESERVED_FIELD_KEYS or normalized_name in _TEXT_ONLY_FIELD_KEYS:
            return
        if _is_scalar_value(value):
            if normalized_name not in field_map:
                field_map[normalized_name] = value
            return
        if isinstance(value, list) and _looks_like_label_value_rows(value):
            for row in value:
                add_field(row.get("label") or row.get("field_name") or row.get("name"), row.get("value"))
            return
        text_value = _value_to_text(value)
        if text_value:
            text_lines.append(f"{normalized_name}: {text_value}")

    structured_fields = structured.get("fields")
    if isinstance(structured_fields, dict):
        for key, value in structured_fields.items():
            if key == "fields" and isinstance(value, list) and _looks_like_label_value_rows(value):
                for row in value:
                    add_field(row.get("label") or row.get("field_name") or row.get("name"), row.get("value"))
                continue
            add_field(key, value)

    for item in extraction_results or []:
        name = item.get("field_name")
        value = item.get("field_value")
        if not name or name.startswith("_") or name in _RESERVED_FIELD_KEYS:
            continue
        add_field(name, value)

    for key, value in structured.items():
        if key in _RESERVED_FIELD_KEYS or key == "fields":
            continue
        if key in _TEXT_ONLY_FIELD_KEYS:
            text_value = _value_to_text(value)
            if text_value:
                text_lines.append(text_value)
            continue
        if _is_scalar_value(value):
            add_field(key, value)
        elif isinstance(value, list) and _looks_like_label_value_rows(value):
            for row in value:
                add_field(row.get("label") or row.get("field_name") or row.get("name"), row.get("value"))
        else:
            text_value = _value_to_text(value)
            if text_value:
                text_lines.append(text_value)

    deduped_text_lines: list[str] = []
    seen_text_lines: set[str] = set()
    for line in text_lines:
        normalized_line = str(line).strip()
        if not normalized_line or normalized_line in seen_text_lines:
            continue
        seen_text_lines.add(normalized_line)
        deduped_text_lines.append(normalized_line)

    return field_map, deduped_text_lines


def _field_rows(field_map: dict[str, Any], extraction_results: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    confidence_by_name: dict[str, Any] = {}
    page_by_name: dict[str, Any] = {}

    for item in extraction_results or []:
        name = item.get("field_name")
        if not name or name.startswith("_") or name in _RESERVED_FIELD_KEYS:
            continue
        if name not in confidence_by_name:
            confidence_by_name[name] = item.get("confidence") if item.get("confidence") is not None else item.get("confidence_score")
        if name not in page_by_name and item.get("page_number") is not None:
            page_by_name[name] = item.get("page_number")

    rows: list[dict[str, Any]] = []
    for name, value in field_map.items():
        rows.append(
            {
                "field_name": name,
                "field_value": value,
                "confidence": confidence_by_name.get(name),
                "page_number": page_by_name.get(name),
            }
        )
    return rows


def _is_scalar_value(value: Any) -> bool:
    return value is None or isinstance(value, (str, int, float, bool))


def _looks_like_label_value_rows(value: Any) -> bool:
    if not isinstance(value, list) or not value:
        return False
    return all(isinstance(item, dict) and ("label" in item or "field_name" in item or "name" in item) and "value" in item for item in value)


def _value_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float, bool)):
        return str(value)
    if isinstance(value, list):
        parts = []
        for item in value:
            item_text = _value_to_text(item)
            if item_text:
                parts.append(item_text)
        return "\n".join(parts).strip()
    if isinstance(value, dict):
        if "label" in value and "value" in value and len(value) <= 3:
            label = str(value.get("label") or "").strip()
            item_value = value.get("value")
            item_text = _value_to_text(item_value)
            if label and item_text:
                return f"{label}: {item_text}"
            if label:
                return label
            return item_text

        parts = []
        for key, item_value in value.items():
            item_text = _value_to_text(item_value)
            if item_text:
                parts.append(f"{key}: {item_text}")
        return "\n".join(parts).strip()
    return str(value).strip()


def _normalize_tables(structured: dict[str, Any], tables: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    source_tables = structured.get("tables") or structured.get("_tables") or (tables or [])
    normalized_tables: list[dict[str, Any]] = []

    for index, table in enumerate(source_tables):
        if not isinstance(table, dict):
            continue

        headers = list(table.get("headers") or [])
        rows = table.get("rows") or []
        if not headers and rows and all(isinstance(row, dict) for row in rows):
            headers = []
            for row in rows:
                for key in row.keys():
                    if key not in headers:
                        headers.append(key)
            rows = [[row.get(header) for header in headers] for row in rows]

        normalized_tables.append(
            {
                "table_index": table.get("table_index", index),
                "page_number": table.get("page_number"),
                "headers": headers,
                "rows": rows,
            }
        )

    return normalized_tables


def _ocr_text(structured: dict[str, Any]) -> str:
    return str((structured.get("ocr") or {}).get("full_text") or "").strip()


def _infer_export_document_type(document: Any, structured: dict[str, Any], extraction_results: list[dict[str, Any]]) -> str:
    doc_type = (structured.get("classification") or {}).get("type") or getattr(document, "doc_type", None) or structured.get("document_type") or "unknown"
    doc_type = str(doc_type).lower()
    text = _ocr_text(structured).lower()
    field_map = _field_lookup(structured, extraction_results)

    invoice_signals = [
        "invoice",
        "invoice number",
        "purchase order",
        "subtotal",
        "balance due",
        "paid to date",
        "amount due",
    ]
    if doc_type == "invoice" or any(signal in text for signal in invoice_signals) or any(name in field_map for name in ("invoice_number", "invoice_date", "due_date", "subtotal", "total")):
        return "invoice"
    # Receipt detection
    receipt_signals = ["receipt", "amount paid", "paid at", "payment received", "merchant"]
    if doc_type == "receipt" or any(signal in text for signal in receipt_signals) or any(name in field_map for name in ("receipt_number", "total_amount", "merchant", "store_name")):
        return "receipt"
    # Business card detection
    bc_signals = ["business card", "contact", "mobile", "email", "phone", "linkedin"]
    if doc_type in ("business_card", "business card") or any(signal in text for signal in bc_signals) or any(name in field_map for name in ("name", "phone", "email", "company")):
        return "business_card"
    # Bank statement detection
    bs_signals = ["statement", "account summary", "opening balance", "closing balance", "transaction", "account number"]
    if doc_type == "bank_statement" or any(signal in text for signal in bs_signals) or any(name in field_map for name in ("account_number", "opening_balance", "closing_balance", "transactions")):
        return "bank_statement"
    return doc_type or "unknown"


def _heuristic_receipt_export(document: Any, structured: dict[str, Any], extraction_results: list[dict[str, Any]], tables: list[dict[str, Any]]) -> dict[str, Any]:
    text = _ocr_text(structured)
    field_map = _field_lookup(structured, extraction_results)
    merchant = field_map.get("merchant") or field_map.get("store_name") or _guess_party_name(text)
    date = field_map.get("date") or ( _guess_invoice_dates(text)[0] if _guess_invoice_dates(text) else None )
    total = field_map.get("total") or field_map.get("amount")
    currency = _guess_currency(text) or "USD"
    return {
        "document_type": "receipt",
        "merchant": merchant,
        "receipt_number": field_map.get("receipt_number") or _guess_invoice_number(text, field_map),
        "date": _normalize_date_value(date),
        "currency": currency,
        "total_amount": _normalize_money(total),
        "line_items": _table_rows_to_items(tables),
        "raw_text_source": {
            "ocr_engine": "EasyOCR",
            "vision_model": f"{Config.LLM_PROVIDER}/{Config.LLM_MODEL}",
            "extraction_timestamp": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        },
    }


def _build_receipt_export_payload(document: Any, structured: dict[str, Any], extraction_results: list[dict[str, Any]], tables: list[dict[str, Any]]) -> dict:
    text = _ocr_text(structured)
    field_map = _field_lookup(structured, extraction_results)
    prompt = (
        "Convert OCR text into a clean receipt JSON. Return ONLY valid JSON with this shape: "
        "{\"document_type\":\"receipt\",\"merchant\":null,\"receipt_number\":null,\"date\":null,\"currency\":null,\"total_amount\":null,\"line_items\":[],\"raw_text_source\":{}}. "
        "Normalize dates to YYYY-MM-DD and currency to numbers. Infer conservatively. "
        f"OCR TEXT:\n{text}\n\nFIELD MAP:\n{json.dumps(field_map, ensure_ascii=False, indent=2)}\n\nTABLES:\n{json.dumps(tables, ensure_ascii=False, indent=2)}"
    )
    llm_payload = None
    try:
        raw = get_llm_client()._call_text(prompt)
        llm_payload = _extract_json(raw)
        if not isinstance(llm_payload, dict):
            llm_payload = None
    except Exception:
        llm_payload = None

    fallback = _heuristic_receipt_export(document, structured, extraction_results, tables)
    if not llm_payload:
        return fallback

    merged = {**fallback, **{k: v for k, v in llm_payload.items() if v not in (None, "")}}
    if merged.get("date"):
        merged["date"] = _normalize_date_value(merged.get("date"))
    if merged.get("total_amount"):
        merged["total_amount"] = _normalize_money(merged.get("total_amount"))
    return merged


def _heuristic_business_card_export(document: Any, structured: dict[str, Any], extraction_results: list[dict[str, Any]], tables: list[dict[str, Any]]) -> dict[str, Any]:
    text = _ocr_text(structured)
    field_map = _field_lookup(structured, extraction_results)
    name = field_map.get("name") or _guess_party_name(text)
    company = field_map.get("company") or field_map.get("organisation") or field_map.get("organization")
    email = field_map.get("email")
    phone = field_map.get("phone")
    title = field_map.get("title")
    address = _guess_address(text)
    return {
        "document_type": "business_card",
        "name": name,
        "title": title,
        "company": company,
        "email": email,
        "phone": phone,
        "address": address,
        "raw_text_source": {
            "ocr_engine": "EasyOCR",
            "vision_model": f"{Config.LLM_PROVIDER}/{Config.LLM_MODEL}",
            "extraction_timestamp": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        },
    }


def _build_business_card_export_payload(document: Any, structured: dict[str, Any], extraction_results: list[dict[str, Any]], tables: list[dict[str, Any]]) -> dict:
    text = _ocr_text(structured)
    field_map = _field_lookup(structured, extraction_results)
    prompt = (
        "Convert OCR text into a clean business card JSON. Return ONLY valid JSON with this shape: "
        "{\"document_type\":\"business_card\",\"name\":null,\"title\":null,\"company\":null,\"email\":null,\"phone\":null,\"address\":{}}. "
        "Normalize emails and phone numbers. Infer conservatively. "
        f"OCR TEXT:\n{text}\n\nFIELD MAP:\n{json.dumps(field_map, ensure_ascii=False, indent=2)}"
    )
    llm_payload = None
    try:
        raw = get_llm_client()._call_text(prompt)
        llm_payload = _extract_json(raw)
        if not isinstance(llm_payload, dict):
            llm_payload = None
    except Exception:
        llm_payload = None

    fallback = _heuristic_business_card_export(document, structured, extraction_results, tables)
    if not llm_payload:
        return fallback

    merged = {**fallback, **{k: v for k, v in llm_payload.items() if v not in (None, "")}}
    return merged


def _heuristic_bank_statement_export(document: Any, structured: dict[str, Any], extraction_results: list[dict[str, Any]], tables: list[dict[str, Any]]) -> dict[str, Any]:
    text = _ocr_text(structured)
    field_map = _field_lookup(structured, extraction_results)
    account_number = field_map.get("account_number")
    opening = field_map.get("opening_balance")
    closing = field_map.get("closing_balance")
    currency = _guess_currency(text) or "USD"
    transactions = []
    # convert first table into transactions if present
    if tables:
        first = tables[0]
        headers = [str(h).strip().lower() for h in (first.get("headers") or [])]
        for row in (first.get("rows") or []):
            values = list(row) if isinstance(row, list) else [row.get(h) for h in headers]
            tx = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
            transactions.append(tx)

    return {
        "document_type": "bank_statement",
        "account_number": account_number,
        "currency": currency,
        "opening_balance": _normalize_money(opening),
        "closing_balance": _normalize_money(closing),
        "transactions": transactions,
        "raw_text_source": {
            "ocr_engine": "EasyOCR",
            "vision_model": f"{Config.LLM_PROVIDER}/{Config.LLM_MODEL}",
            "extraction_timestamp": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        },
    }


def _build_bank_statement_export_payload(document: Any, structured: dict[str, Any], extraction_results: list[dict[str, Any]], tables: list[dict[str, Any]]) -> dict:
    text = _ocr_text(structured)
    field_map = _field_lookup(structured, extraction_results)
    prompt = (
        "Convert OCR text into a clean bank statement JSON. Return ONLY valid JSON with this shape: "
        "{\"document_type\":\"bank_statement\",\"account_number\":null,\"currency\":null,\"opening_balance\":null,\"closing_balance\":null,\"transactions\":[]}. "
        "Normalize currency amounts and dates. Infer conservatively. "
        f"OCR TEXT:\n{text}\n\nFIELD MAP:\n{json.dumps(field_map, ensure_ascii=False, indent=2)}\n\nTABLES:\n{json.dumps(tables, ensure_ascii=False, indent=2)}"
    )
    llm_payload = None
    try:
        raw = get_llm_client()._call_text(prompt)
        llm_payload = _extract_json(raw)
        if not isinstance(llm_payload, dict):
            llm_payload = None
    except Exception:
        llm_payload = None

    fallback = _heuristic_bank_statement_export(document, structured, extraction_results, tables)
    if not llm_payload:
        return fallback

    merged = {**fallback, **{k: v for k, v in llm_payload.items() if v not in (None, "")}}
    return merged


def _normalize_money(value: Any) -> float | None:
    return normalize_currency(value)


def _normalize_date_value(value: Any) -> str | None:
    if value is None:
        return None
    normalized = normalize_date(str(value))
    return normalized or str(value).strip()


def _guess_invoice_number(text: str, field_map: dict[str, Any]) -> str | None:
    for key in ("invoice_number", "invoice no", "invoice_no", "invoice #"):
        value = field_map.get(key)
        if value:
            return str(value).strip()
    patterns = [
        r"invoice\s*(?:no\.?|number|#|_|:)\s*([A-Z0-9\-/]+)",
        r"\binv[-_ ]?([A-Z0-9\-/]+)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return None


def _guess_purchase_order_number(text: str, invoice_number: str | None) -> str | None:
    patterns = [
        r"purchase\s*order\s*(?:number|no\.?|#|:)\s*([A-Z0-9\-/]+)",
        r"po\s*(?:number|no\.?|#|:)\s*([A-Z0-9\-/]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return invoice_number


def _guess_currency(text: str) -> str | None:
    if re.search(r"\bUSD\b|\$", text, re.IGNORECASE):
        return "USD"
    if re.search(r"\bEUR\b|€", text, re.IGNORECASE):
        return "EUR"
    if re.search(r"\bGBP\b|£", text, re.IGNORECASE):
        return "GBP"
    return None


def _guess_invoice_dates(text: str) -> list[str]:
    dates = []
    for pattern in [r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+\d{1,2},?\s+\d{4}\b", r"\b\d{1,2}\s+(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+\d{4}\b", r"\b\d{4}-\d{2}-\d{2}\b"]:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            normalized = _normalize_date_value(match.group(0))
            if normalized and normalized not in dates:
                dates.append(normalized)
    return dates


def _guess_email(text: str) -> str | None:
    match = re.search(r"\b[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}\b", text)
    return match.group(0).strip() if match else None


def _normalize_ifsc(value: str | None) -> str | None:
    if not value:
        return None
    candidate = str(value).strip().upper().replace(" ", "")
    candidate = candidate.replace("O", "0").replace("Q", "0")
    return candidate


def _guess_bank_details(text: str) -> dict[str, Any]:
    account_holder = None
    account_number = None
    ifsc_code = None
    swift_code = None
    account_type = None

    match = re.search(r"(?:invoice\s*name|name)[:_\-\s]*([A-Z][A-Za-z .'-]+?)(?:\s+account\s+no|\s+date|\s+invoice|$)", text, re.IGNORECASE)
    if match:
        account_holder = match.group(1).strip().rstrip(".,")

    match = re.search(r"account\s*no(?:\.|:)?\s*([0-9]{6,})", text, re.IGNORECASE)
    if match:
        account_number = match.group(1).strip()

    match = re.search(r"ifsc(?:\s*code)?\s*[:\-]?\s*([A-Z0-9]{8,})", text, re.IGNORECASE)
    if match:
        ifsc_code = _normalize_ifsc(match.group(1))

    match = re.search(r"swift\s*code\s*[:\-]?\s*([A-Z0-9]{6,11})", text, re.IGNORECASE)
    if match:
        swift_code = match.group(1).strip().upper()

    match = re.search(r"account\s*type\s*[:\-]?\s*([A-Za-z ]+)" , text, re.IGNORECASE)
    if match:
        account_type = match.group(1).strip().rstrip(".,")

    return {
        "account_holder": account_holder,
        "account_number": account_number,
        "ifsc_code": ifsc_code,
        "swift_code": swift_code,
        "account_type": account_type,
    }


def _guess_party_name(text: str, preferred: str | None = None) -> str | None:
    if preferred:
        return str(preferred).strip()
    candidates = re.findall(r"\b[A-Z][a-z]+(?:\s+[A-Z][a-z]+)+\b", text)
    return candidates[0] if candidates else None


def _guess_address(text: str) -> dict[str, Any]:
    postal_match = re.search(r"(\d{6})", text)
    postal_code = postal_match.group(1) if postal_match else None
    street = city = state = country = None

    if postal_code:
        before = text[: postal_match.start()].strip()
        chunks = [chunk.strip() for chunk in re.split(r"[,\n]", before) if chunk.strip()]
        if chunks:
            street = chunks[-1]
        if len(chunks) >= 2:
            city = chunks[-2]
        if len(chunks) >= 3:
            state = chunks[-3]

    country_match = re.search(r"\b(India|United States|USA|UAE|United Kingdom|UK|Canada|Australia)\b", text, re.IGNORECASE)
    if country_match:
        country = country_match.group(1)

    return {
        "street": street,
        "city": city,
        "state": state,
        "postal_code": postal_code,
        "country": country,
    }


def _table_rows_to_items(tables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not tables:
        return []
    first = tables[0] or {}
    headers = [str(item).strip().lower() for item in (first.get("headers") or [])]
    rows = first.get("rows") or []
    items: list[dict[str, Any]] = []
    for row in rows:
        values = list(row) if isinstance(row, list) else [row.get(header, "") for header in headers]
        item: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            value = values[idx] if idx < len(values) else None
            if header in ("description", "item", "product", "service"):
                item["description"] = value
            elif header in ("quantity", "qty"):
                item["quantity"] = normalize_currency(value)
            elif header in ("unit_price", "unit price", "rate", "price"):
                item["unit_price"] = normalize_currency(value)
            elif header in ("amount", "total"):
                item["amount"] = normalize_currency(value)
        if item:
            items.append({k: v for k, v in item.items() if v is not None})
    return items


def _confidence_scores(invoice_number: Any, dates: list[str], totals: dict[str, Any], bank_details: dict[str, Any], items: list[dict[str, Any]]) -> dict[str, float]:
    return {
        "invoice_number": 0.99 if invoice_number else 0.3,
        "dates": 0.98 if dates else 0.3,
        "amounts": 0.99 if totals.get("total_amount") is not None else 0.3,
        "bank_details": 0.96 if any(bank_details.values()) else 0.3,
        "line_items": 0.97 if items else 0.3,
    }


def _heuristic_invoice_export(document: Any, structured: dict[str, Any], extraction_results: list[dict[str, Any]], tables: list[dict[str, Any]]) -> dict[str, Any]:
    text = _ocr_text(structured)
    field_map = _field_lookup(structured, extraction_results)

    invoice_number = _guess_invoice_number(text, field_map)
    purchase_order_number = _guess_purchase_order_number(text, invoice_number)
    dates = _guess_invoice_dates(text)
    invoice_date = dates[0] if dates else None
    due_date = dates[1] if len(dates) > 1 else invoice_date
    currency = _guess_currency(text) or "USD"

    seller_name = _guess_party_name(text, field_map.get("vendor_name") or field_map.get("seller_name") or field_map.get("name"))
    seller_address = _guess_address(text)
    seller_email = _guess_email(text)
    seller_bank_details = _guess_bank_details(text)

    buyer_name = None
    buyer_match = re.search(r"invoice\s*due\s*to\s*(.+?)(?:\s+purchase\s*order|\s+po\s*number|\s+description|$)", text, re.IGNORECASE)
    if buyer_match:
        raw_buyer = buyer_match.group(1).strip()
        name_match = re.search(r"([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)", raw_buyer)
        if name_match:
            buyer_name = name_match.group(1).strip()
    if not buyer_name:
        candidates = re.findall(r"\b[A-Z][a-z]+(?:\s+[A-Z][a-z]+)+\b", text)
        if seller_name and seller_name in candidates:
            candidates = [item for item in candidates if item != seller_name]
        buyer_name = candidates[0] if candidates else None

    # totals
    subtotal = None
    total = None
    paid_to_date = None
    balance_due = None
    subtotal_match = re.search(r"sub\s*total\s*([A-Z]{0,3}\s*)?([0-9,]+(?:\.[0-9]+)?)", text, re.IGNORECASE)
    if subtotal_match:
        subtotal = _normalize_money(subtotal_match.group(2))
    total_match = re.search(r"\btotal\s*(?:[A-Z]{3}\s*)?([0-9,]+(?:\.[0-9]+)?)", text, re.IGNORECASE)
    if total_match:
        total = _normalize_money(total_match.group(1))
    paid_match = re.search(r"paid\s*to\s*date\s*(?:[A-Z]{3}\s*)?([0-9,]+(?:\.[0-9]+)?)", text, re.IGNORECASE)
    if paid_match:
        paid_to_date = _normalize_money(paid_match.group(1))
    balance_match = re.search(r"balance\s*(?:due|\))?\s*(?:[A-Z]{3}\s*)?([0-9,]+(?:\.[0-9]+)?)", text, re.IGNORECASE)
    if balance_match:
        balance_due = _normalize_money(balance_match.group(1))

    if subtotal is None and total is not None:
        subtotal = total
    if paid_to_date is None:
        paid_to_date = 0.0 if balance_due or total else None
    if balance_due is None and total is not None and paid_to_date is not None:
        balance_due = max((total or 0.0) - paid_to_date, 0.0)
    if total is None and subtotal is not None:
        total = subtotal

    note_match = re.search(r"invoice\s*note\s*(.*?)(?:many\s+thanks|regards|best\s+regards|$)", text, re.IGNORECASE)
    invoice_note = None
    if note_match:
        invoice_note = re.sub(r"\s+", " ", note_match.group(1)).strip().rstrip("._")

    items = _table_rows_to_items(tables)
    if not items and subtotal is not None and total is not None:
        description_match = re.search(r"worked\s+on\s+(.+?)\s+modif", text, re.IGNORECASE)
        items = [
            {
                "description": description_match.group(1).strip() if description_match else None,
                "quantity": None,
                "unit_price": None,
                "amount": total,
            }
        ]
        items = [{k: v for k, v in item.items() if v is not None} for item in items]

    payment_status = "unpaid"
    if balance_due is not None:
        if balance_due <= 0:
            payment_status = "paid"
        elif paid_to_date and paid_to_date > 0:
            payment_status = "partial"

    export = {
        "document_type": "invoice",
        "invoice_details": {
            "invoice_number": invoice_number,
            "purchase_order_number": purchase_order_number,
            "invoice_date": invoice_date,
            "due_date": due_date,
            "currency": currency,
        },
        "seller": {
            "name": seller_name,
            "address": seller_address,
            "email": seller_email,
            "bank_details": seller_bank_details,
        },
        "buyer": {
            "name": buyer_name,
            "country": _guess_address(text).get("country") if buyer_name else None,
        },
        "items": items,
        "totals": {
            "subtotal": subtotal,
            "paid_to_date": paid_to_date,
            "balance_due": balance_due,
            "total_amount": total,
        },
        "payment_status": payment_status,
        "invoice_note": invoice_note,
        "confidence_scores": _confidence_scores(invoice_number, [d for d in [invoice_date, due_date] if d], {"total_amount": total}, seller_bank_details, items),
        "raw_text_source": {
            "ocr_engine": "EasyOCR",
            "vision_model": f"{Config.LLM_PROVIDER}/{Config.LLM_MODEL}",
            "extraction_timestamp": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        },
    }

    return export


def _build_invoice_export_payload(document: Any, structured: dict[str, Any], extraction_results: list[dict[str, Any]], tables: list[dict[str, Any]]) -> dict:
    text = _ocr_text(structured)
    field_map = _field_lookup(structured, extraction_results)

    prompt = (
        "You are converting OCR text into a clean invoice JSON export. "
        "Return ONLY valid JSON and do not include any OCR internals, bbox data, or explanations. "
        "Correct OCR mistakes contextually, normalize dates to YYYY-MM-DD, normalize currency to numbers, and fill missing values with null. "
        "Use this exact JSON shape: "
        "{\"document_type\":\"invoice\",\"invoice_details\":{\"invoice_number\":null,\"purchase_order_number\":null,\"invoice_date\":null,\"due_date\":null,\"currency\":null},"
        "\"seller\":{\"name\":null,\"address\":{\"street\":null,\"city\":null,\"state\":null,\"postal_code\":null,\"country\":null},\"email\":null,\"bank_details\":{\"account_holder\":null,\"account_number\":null,\"ifsc_code\":null,\"swift_code\":null,\"account_type\":null}},"
        "\"buyer\":{\"name\":null,\"country\":null},\"items\":[],\"totals\":{\"subtotal\":null,\"paid_to_date\":null,\"balance_due\":null,\"total_amount\":null},"
        "\"payment_status\":null,\"invoice_note\":null,\"confidence_scores\":{\"invoice_number\":0,\"dates\":0,\"amounts\":0,\"bank_details\":0,\"line_items\":0},"
        "\"raw_text_source\":{\"ocr_engine\":\"EasyOCR\",\"vision_model\":\"\",\"extraction_timestamp\":\"\"}}. "
        "If some data is not explicit, infer it conservatively from OCR text and tables. "
        f"OCR TEXT:\n{text}\n\n"
        f"FIELD MAP:\n{json.dumps(field_map, ensure_ascii=False, indent=2)}\n\n"
        f"TABLES:\n{json.dumps(tables, ensure_ascii=False, indent=2)}"
    )

    llm_payload: dict[str, Any] | None = None
    try:
        raw = get_llm_client()._call_text(prompt)
        llm_payload = _extract_json(raw)
        if not isinstance(llm_payload, dict):
            llm_payload = None
    except Exception:
        llm_payload = None

    fallback = _heuristic_invoice_export(document, structured, extraction_results, tables)

    if not llm_payload:
        return fallback

    # Merge LLM output over heuristic fallback, keeping required schema stable.
    merged = fallback
    for key, value in llm_payload.items():
        if key in merged and isinstance(merged[key], dict) and isinstance(value, dict):
            merged[key] = {**merged[key], **value}
        elif key in merged and isinstance(merged[key], list) and isinstance(value, list):
            merged[key] = value or merged[key]
        elif value not in (None, ""):
            merged[key] = value

    # Final cleanup for required primitives.
    if merged.get("invoice_details"):
        merged["invoice_details"]["invoice_date"] = _normalize_date_value(merged["invoice_details"].get("invoice_date"))
        merged["invoice_details"]["due_date"] = _normalize_date_value(merged["invoice_details"].get("due_date"))
    if merged.get("totals"):
        for key in ("subtotal", "paid_to_date", "balance_due", "total_amount"):
            merged["totals"][key] = _normalize_money(merged["totals"].get(key))

    return merged


def export_batch_zip(files: list[str], output_zip_path: str) -> str:
    path = Path(output_zip_path)
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for file_path in files:
            archive.write(file_path, arcname=Path(file_path).name)
    return str(path)
